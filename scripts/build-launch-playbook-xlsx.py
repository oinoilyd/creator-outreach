#!/usr/bin/env python3
"""
Build the launch playbook as an .xlsx file.

Output: docs/launch-playbook.xlsx — proper Excel spreadsheet with
styled headers, owner color-coding, and a dropdown for Status so
Dylan + Ryan can mark items Done as they complete.

Layout (single sheet):
  TOP CONTEXT      — strategic frame: critical path + video strategy
  ACTION TABLE     — 17 rows of who-does-what (Dylan / Ryan / Joint)
  BOTTOM NOTES     — discount design, testimonials, video split detail,
                     conversion targets, deferred items

Re-run any time the playbook content changes:
    python3 scripts/build-launch-playbook-xlsx.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_PATH = "docs/launch-playbook.xlsx"

# ── Row schema ──────────────────────────────────────────────────────
# (number, owner, action, why_it_matters, time_estimate)
ROWS = [
    # Dylan's actions
    (1, "Dylan", "Build product demo video (3-5 min)",
     "Walk through AI fit score, 5-platform search, outreach send, CRM, custom analytics. Practical, not narrative. Ryan links this in warm intros so leads can see the product.",
     "2-4 hrs"),
    (2, "Dylan", "Write Ryan's warm-intro email template",
     "First impression for all 100 leads. Casual + credible + clear ask.",
     "15 min"),
    (3, "Dylan", "Write follow-up email sequence (first reply + 2-3 follow-ups)",
     "Conversion determinant. Most leads need 2-3 touches before they bite.",
     "30 min"),
    (4, "Dylan", "Set up early-customer discount in Stripe",
     "First 10 customers get $20/mo locked forever OR 2 months free. Urgency + reward for early adopters. Stripe coupons handle it.",
     "15 min"),
    (5, "Dylan", "Walk through onboarding as a new user",
     "Incognito → signup → first search → first outreach. Fix anything janky. First 5 minutes determines 5/10 vs 8/10 conversion.",
     "30 min"),
    (6, "Dylan", "Polish /pricing page FAQ",
     "Anticipate top 5 objections + answer inline. Where leads decide yes/no.",
     "30 min"),
    (7, "Dylan", "Plan testimonial collection",
     "What's the moment you ask each customer for a quote? First 5 testimonials = next 100 customers.",
     "15 min"),

    # Ryan's actions
    (8, "Ryan", "Reply with Gaynor Media LLC info",
     "State, county, full registered address, EIN. Unblocks Stripe legal-name fields + legal-doc placeholders.",
     "1 min"),
    (9, "Ryan", "Review employment contract for non-compete / non-solicit",
     "Flag any clauses that restrict working on a side venture. Identify accounts in the 100-lead pipeline that are 'off-limits' (current employer's customers).",
     "20 min"),
    (10, "Ryan", "Curate the 100 leads",
     "Segment into 'safe to intro' vs 'skip' based on contract review (#9). Build the actual list with email + context per lead.",
     "30-60 min"),
    (11, "Ryan", "Produce founder story video (2-4 min)",
     "Ryan tells his story — industry experience, why this product, what it means to him — ending with the product as the resolution. Ryan's lane as a video creator/editor. Sells the WHY; product demo (#1) sells the WHAT. Both linked in the warm intro.",
     "1-2 days (Ryan's editing time)"),
    (12, "Ryan", "Send the warm intros",
     "Use the intro template (#2) to the curated leads (#10), batch by batch so Dylan can keep up with inbound. This is where the 5-10 first customers come from.",
     "1-2 hrs across launch week"),

    # Joint actions
    (13, "Joint", "Dylan follows up + closes",
     "Reply to interested leads, book calls if needed, send checkout link, handle questions.",
     "Async, ongoing during launch"),
    (14, "Joint", "Daily 5-min sync during launch week",
     "What's converting, what's not, what to tweak. Async OK (Slack / text). Most learnings happen in the first 10 leads.",
     "5 min/day for ~2 weeks"),
]

# ── TOP CONTEXT — strategic frame above the action table ──────────
TOP_CONTEXT = [
    ("CREATOR OUTREACH — LAUNCH PLAYBOOK", "title"),
    ("Target: First 5-10 paying customers via Ryan's 100 warm leads. 5% = $250 MRR baseline. 10% = $500 MRR strong launch.", "subtitle"),
    ("", "blank"),
    ("CRITICAL PATH:  Ryan replies w/ LLC info (#8)  →  [Dylan polishes product + writes emails (#1-7)]  +  [Ryan reviews contract, curates leads, films founder video (#9-11)]  →  Ryan sends warm intros (#12)  →  Dylan closes (#13)  →  Daily sync during launch (#14)", "callout"),
    ("", "blank"),
    ("VIDEO STRATEGY — two videos, two functions, BOTH linked in the warm intro:", "header"),
    ("• FOUNDER STORY VIDEO (Ryan, item #13): 2-4 min. Ryan tells his story — industry experience, why this product, what it means to him — ending with the product as the resolution. Ryan is a video creator/editor; this is his lane. Sells the WHY. Builds emotional trust with cold-warm leads.", "bullet"),
    ("• PRODUCT DEMO VIDEO (Dylan, item #1): 3-5 min. Practical walkthrough of AI fit score, 5-platform search, outreach send, CRM, custom analytics. Sells the WHAT. Builds belief in the product itself.", "bullet"),
    ("• ORDER IN INTRO EMAIL: founder video first (hook). 'If you want to see the product itself, here's a 5-min demo.' (link).", "bullet"),
    ("", "blank"),
    ("OWNER COLOR CODE — Dylan = light blue,  Ryan = light yellow,  Joint = light green", "subtitle"),
    ("", "blank"),
]

# ── Build the workbook ─────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Launch Playbook"

# Style definitions
title_font = Font(bold=True, color="FFFFFF", size=16)
title_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
subtitle_font = Font(italic=True, color="374151", size=10)
header_font_top = Font(bold=True, color="111827", size=12)
callout_font = Font(color="1F2937", size=10)
callout_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
bullet_font = Font(color="374151", size=10)
notes_header_font = Font(bold=True, color="111827", size=12)
notes_body_font = Font(color="374151", size=10)

# Top context rendering
current_row = 1
for text, style in TOP_CONTEXT:
    ws.cell(row=current_row, column=1, value=text)
    # Merge across all 7 columns for narrative rows
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
    cell = ws.cell(row=current_row, column=1)
    if style == "title":
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[current_row].height = 35
    elif style == "subtitle":
        cell.font = subtitle_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[current_row].height = 25
    elif style == "header":
        cell.font = header_font_top
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[current_row].height = 25
    elif style == "callout":
        cell.font = callout_font
        cell.fill = callout_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[current_row].height = 45
    elif style == "bullet":
        cell.font = bullet_font
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[current_row].height = 35
    else:  # blank
        ws.row_dimensions[current_row].height = 10
    current_row += 1

# Skip a row before the action table
current_row += 1

# Column headers for the action table
HEADERS = ["#", "Owner", "Action", "Why It Matters", "Time", "Status", "Notes"]
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
header_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
for col_idx, header in enumerate(HEADERS, start=1):
    cell = ws.cell(row=current_row, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
ws.row_dimensions[current_row].height = 30
header_row = current_row
current_row += 1

# Owner color coding
OWNER_FILLS = {
    "Dylan": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),  # light blue
    "Ryan":  PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),  # light yellow
    "Joint": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),  # light green
}

# Data row styling
data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='E5E7EB'),
    right=Side(style='thin', color='E5E7EB'),
    top=Side(style='thin', color='E5E7EB'),
    bottom=Side(style='thin', color='E5E7EB'),
)

table_start_row = current_row
for num, owner, action, why, time_est in ROWS:
    ws.cell(row=current_row, column=1, value=num)
    ws.cell(row=current_row, column=2, value=owner)
    ws.cell(row=current_row, column=3, value=action)
    ws.cell(row=current_row, column=4, value=why)
    ws.cell(row=current_row, column=5, value=time_est)
    ws.cell(row=current_row, column=6, value="Pending")
    ws.cell(row=current_row, column=7, value="")
    # Apply alignment + border to all cells in row
    for col_idx in range(1, len(HEADERS) + 1):
        c = ws.cell(row=current_row, column=col_idx)
        c.alignment = data_alignment
        c.border = thin_border
    # Color-code the Owner cell + the row number cell
    owner_fill = OWNER_FILLS[owner]
    ws.cell(row=current_row, column=1).fill = owner_fill  # #
    ws.cell(row=current_row, column=2).fill = owner_fill  # Owner
    # Bold the action text
    ws.cell(row=current_row, column=3).font = Font(bold=True)
    ws.row_dimensions[current_row].height = 60
    current_row += 1

table_end_row = current_row - 1

# Status dropdown — Pending / In Progress / Done / Blocked
dv = DataValidation(
    type="list",
    formula1='"Pending,In Progress,Done,Blocked"',
    allow_blank=True,
)
dv.add(f"F{table_start_row}:F{table_end_row}")
ws.add_data_validation(dv)

# ── BOTTOM NOTES — open questions, deferred items, video reminder ──
current_row += 2  # spacer

BOTTOM_NOTES = [
    ("NOTES, OPEN QUESTIONS & DEFERRED ITEMS", "title_small"),
    ("", "blank"),
    ("VIDEO STRATEGY (re-statement for visibility):", "header"),
    ("Ryan's founder video and Dylan's product demo serve different functions and are NOT redundant. Founder video = emotional trust-builder, 2-4 min, Ryan-produced. Demo video = practical walkthrough, 3-5 min, Dylan-produced. Both must exist before the warm-intro push. Link both in every intro email — founder video first as the hook.", "body"),
    ("", "blank"),
    ("DISCOUNT DESIGN (item #4) — decide before launch:", "header"),
    ("• Option A: $20/mo locked forever for first 10 customers (lifetime founder pricing). Strongest signal of value. Hardest to walk back later.", "body"),
    ("• Option B: 2 months free at $50/mo. Easier to upsell. Less margin compression.", "body"),
    ("• Option C: 50% off first 6 months. Middle ground. Standard SaaS launch playbook.", "body"),
    ("", "blank"),
    ("TESTIMONIAL COLLECTION (item #7) — ideas:", "header"),
    ("• Founder DM after first successful outreach campaign: 'got a sec for a 1-line quote?'", "body"),
    ("• In-app banner asking after 14 days of use.", "body"),
    ("• During the daily sync — flag a happy customer, both founders prompt them in parallel.", "body"),
    ("", "blank"),
    ("CONVERSION TARGETS — calibration for the 100-lead push:", "header"),
    ("• 5% conversion = 5 paying customers = $250 MRR = $3,000 ARR — good baseline.", "body"),
    ("• 10% conversion = 10 paying customers = $500 MRR = $6,000 ARR — strong launch.", "body"),
    ("• 15%+ conversion — strong founder-led sales signal. Hire-decision territory.", "body"),
    ("", "blank"),
    ("DEFERRED ITEMS — not blocking launch:", "header"),
    ("• Insurance quote (Vouch / Embroker) — $1-3K/yr, cyber + E&O. Get once first customers exist.", "body"),
    ("• Supabase Pro upgrade ($25/mo) — before first paying customer (daily backups).", "body"),
    ("• Stripe trial-end reminder email toggle — 30 sec, do whenever.", "body"),
    ("• Namecheap LLC card swap — at next renewal (~11 months out).", "body"),
    ("", "blank"),
    ("REGENERATE: edit ROWS in scripts/build-launch-playbook-xlsx.py and run `python3 scripts/build-launch-playbook-xlsx.py`", "footnote"),
]

for text, style in BOTTOM_NOTES:
    ws.cell(row=current_row, column=1, value=text)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
    cell = ws.cell(row=current_row, column=1)
    if style == "title_small":
        cell.font = Font(bold=True, color="FFFFFF", size=13)
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[current_row].height = 30
    elif style == "header":
        cell.font = notes_header_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[current_row].height = 22
    elif style == "body":
        cell.font = notes_body_font
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[current_row].height = 30
    elif style == "footnote":
        cell.font = Font(italic=True, color="6B7280", size=9)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[current_row].height = 25
    else:  # blank
        ws.row_dimensions[current_row].height = 8
    current_row += 1

# Column widths (tuned for readability)
COLUMN_WIDTHS = {
    "A": 5,    # #
    "B": 10,   # Owner
    "C": 45,   # Action
    "D": 65,   # Why It Matters
    "E": 22,   # Time
    "F": 14,   # Status
    "G": 30,   # Notes
}
for col_letter, width in COLUMN_WIDTHS.items():
    ws.column_dimensions[col_letter].width = width

# No freeze panes — the bottom notes section is more readable when
# the whole document scrolls naturally together. Earlier version
# froze the top context (rows 1-14) which made the bottom feel
# cramped during scrolling.

# ── Save ───────────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)
print(f"✓ Wrote {OUTPUT_PATH}")
print(f"  - Top context: rows 1-{table_start_row - 2}")
print(f"  - Action table: rows {header_row}-{table_end_row}")
print(f"  - Bottom notes: rows {table_end_row + 3}+")
